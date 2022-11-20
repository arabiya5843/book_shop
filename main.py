import  traceback
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from models import Book_shop , After , Literature , Genre , Publishing , User

class Login_window(QDialog):

    def __init__(self):
        super().__init__()

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setGeometry(100, 100, 400, 200)
        self.setWindowTitle("Kirish")
        self.setStyleSheet("background-color: #dd99ff;")

        #Login
        self.login = QLabel(self)
        self.login.setText("Login")
        self.login.setFont(QFont("Terminus", 70))
        self.login.move(800, 280)

        self.lb_login = QLabel(self)
        self.lb_login.setText("Login:")
        self.lb_login.setFont(QFont("Terminus",30))
        self.lb_login.move(652 , 460)

        self.le_login = QLineEdit(self)
        self.le_login.setFont(QFont("Terminus",14))
        self.le_login.setGeometry(890 , 481 , 300 , 40)
        self.le_login.setStyleSheet("QLineEdit" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        #Password
        self.lb_password = QLabel(self)
        self.lb_password.setText("Password:")
        self.lb_password.setFont(QFont("Terminus",30))
        self.lb_password.move(652 , 530)

        self.le_password = QLineEdit(self)
        self.le_password.setEchoMode(QLineEdit.Password)
        self.le_password.setFont(QFont("Terminus",14))
        self.le_password.setGeometry(890 , 551 , 300 , 40)
        self.le_password.setStyleSheet("QLineEdit" "{"
                                       "background : white;"
                                       "border : 8px solid;"
                                       "border-top-color : grey;"
                                       "border-left-color : grey;"
                                       "border-right-color : #ebebe0;"
                                       "border-bottom-color : #ebebe0;" "}")

        #Button
        self.button = QPushButton(self)
        self.button.setText("Click")
        self.button.setFont(QFont("Terminus" , 15))
        self.button.setGeometry(810 , 640 , 250 , 50)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.button.setStyleSheet(QBTN)
        self.button.clicked.connect(self.on_click_btn_login)

    def on_click_btn_login(self):
        login = self.le_login.text()
        password = self.le_password.text()

        import hashlib
        encoded = hashlib.sha256()
        encoded.update(password.encode("ascii"))

        for item in User.objects():
            if login == item.user_name and encoded.hexdigest() == item.password:
                self.user = item
                self.accept()
                return

        self.msg = QMessageBox(self)
        self.msg.setText("User topilamdi.Iltimos qayta urinib ko'ring!")
        self.msg.setIcon(QMessageBox.Critical)
        self.msg.setWindowTitle("Error⚠️")
        self.msg.setStyleSheet("background-color : white")
        self.msg.show()


class Shop_window(QMainWindow):

    def __init__(self):
        super().__init__()

        self.init_menu()

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setWindowTitle("Book Shop")
        self.setStyleSheet("background-color: #dd99ff;")

        #Name
        self.lb_name = QLabel(self)
        self.lb_name.setText("Name:")
        self.lb_name.setFont(QFont("Normal",20))
        self.lb_name.setGeometry(30 , 40 , 200 , 40)

        self.le_name = QLineEdit(self)
        self.le_name.setFont(QFont("Normal",14))
        self.le_name.setGeometry(220 , 40 , 200 , 40)
        self.le_name.setStyleSheet("QLineEdit" "{"
                                       "background : white;"
                                       "border : 8px solid;"
                                       "border-top-color : grey;"
                                       "border-left-color : grey;"
                                       "border-right-color : #ebebe0;"
                                       "border-bottom-color : #ebebe0;" "}")

        #Price
        self.lb_price = QLabel(self)
        self.lb_price.setText("Price:")
        self.lb_price.setFont(QFont("Normal",20))
        self.lb_price.setGeometry(30 , 100, 200 , 40)

        self.le_price = QLineEdit(self)
        self.le_price.setFont(QFont("Normal",14))
        self.le_price.setGeometry(220 , 100 , 200 , 40)
        self.le_price.setStyleSheet("QLineEdit" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        #Amount
        self.lb_amount = QLabel(self)
        self.lb_amount.setText("Amount:")
        self.lb_amount.setFont(QFont("Normal", 20))
        self.lb_amount.setGeometry(30, 160, 200, 40)

        self.sb_amount = QSpinBox(self)
        self.sb_amount.setMinimum(1)
        self.sb_amount.setFont(QFont("Normal", 14))
        self.sb_amount.setGeometry(220 , 160 , 200 , 40)
        self.sb_amount.setStyleSheet("QSpinBox" "{"
                                     "background : white;"
                                     "border : 8px solid;"
                                     "border-top-color : grey;"
                                     "border-left-color : grey;"
                                     "border-right-color : #ebebe0;"
                                     "border-bottom-color : #ebebe0;" "}")

        #After
        self.lb_after = QLabel(self)
        self.lb_after.setText("After:")
        self.lb_after.setFont(QFont("Normal", 20))
        self.lb_after.setGeometry(30, 220 , 200 , 40)

        self.cb_after = QComboBox(self)
        self.cb_after.setFont(QFont("Normal",14))
        self.cb_after.setGeometry(220 , 220 , 200 , 40)
        self.cb_after.setStyleSheet("QComboBox" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        for item in After.objects():
            self.cb_after.addItem(item.afters_name, userData=item.id)

        #Literature
        self.lb_literature = QLabel(self)
        self.lb_literature.setText("Literature:")
        self.lb_literature.setFont(QFont("Normal", 20))
        self.lb_literature.setGeometry(30 , 280 , 200  , 40)

        self.cb_literature = QComboBox(self)
        self.cb_literature.setFont(QFont("Normal", 14))
        self.cb_literature.setGeometry(220, 280, 200, 40)
        self.cb_literature.setStyleSheet("QComboBox" "{"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        for item in Literature.objects():
            self.cb_literature.addItem(item.folk_literature, userData=item.id)

        #Genre
        self.lb_genre = QLabel(self)
        self.lb_genre.setText("Genre:")
        self.lb_genre.setFont(QFont("Normal", 20))
        self.lb_genre.move(30, 350)

        self.cb_genre = QComboBox(self)
        self.cb_genre.setFont(QFont("Normal", 14))
        self.cb_genre.setGeometry(220, 340, 200, 40)
        self.cb_genre.setStyleSheet("QComboBox" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        for item in Genre.objects():
            self.cb_genre.addItem(item.genre, userData=item.id)

        #Publishing
        self.lb_publishing = QLabel(self)
        self.lb_publishing.setText("Publishing:")
        self.lb_publishing.setFont(QFont("Normal", 20))
        self.lb_publishing.setGeometry(30 , 400 , 200  , 50)

        self.cb_publishing = QComboBox(self)
        self.cb_publishing.setFont(QFont("Normal", 14))
        self.cb_publishing.setGeometry(220, 400, 200, 40)
        self.cb_publishing.setStyleSheet("QComboBox" "{"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        for item in Publishing.objects():
            self.cb_publishing.addItem(item.publishing, userData=item.id)

        #Button
        self.save_btn = QPushButton(self)
        self.save_btn.setText("Save")
        self.save_btn.setFont(QFont("Normal", 12))
        self.save_btn.setGeometry(220 , 460 , 200 , 45)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.save_btn.setStyleSheet(QBTN)
        self.save_btn.clicked.connect(self.on_click_btn_save)

        self.update_btn = QPushButton(self)
        self.update_btn.setText("Update")
        self.update_btn.setFont(QFont("Normal" , 12))
        self.update_btn.setGeometry(220, 520 , 200 , 45)
        QBTN = """
                   QPushButton:active{
                     background-color : #00ff00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   QPushButton:pressed{
                     background-color : #00cc00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   """
        self.update_btn.setStyleSheet(QBTN)
        self.update_btn.clicked.connect(self.on_click_btn_update)

        self.delete_btn = QPushButton(self)
        self.delete_btn.setText("Delete")
        self.delete_btn.setFont(QFont("Normal" , 12))
        self.delete_btn.setGeometry(220, 580 , 200 , 45)
        QBTN = """
                   QPushButton:active{
                     background-color : #00ff00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   QPushButton:pressed{
                     background-color : #00cc00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   """
        self.delete_btn.setStyleSheet(QBTN)
        self.delete_btn.clicked.connect(self.on_click_btn_delete)

        self.report_btn = QPushButton(self)
        self.report_btn.setText("Report")
        self.report_btn.setFont(QFont("Normal", 12))
        self.report_btn.setGeometry(220, 640 , 200 , 45)
        QBTN = """
                   QPushButton:active{
                     background-color : #00ff00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   QPushButton:pressed{
                     background-color : #00cc00;
                     border: 8px solid #00ff00;
                     border-top-color : grey;
                     border-left-color : grey;
                     border-right-color : #ebebe0;
                     border-bottom-color : #ebebe0;
                   }
                   """
        self.report_btn.setStyleSheet(QBTN)
        self.report_btn.clicked.connect(self.on_click_btn_report)

        self.le_search = QLineEdit(self)
        self.le_search.setPlaceholderText("Qidiruv")
        self.le_search.setFont(QFont("Normal", 14))
        self.le_search.setGeometry(220 , 700 , 200 , 45)
        self.le_search.setStyleSheet("QLineEdit" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        self.le_search.textChanged.connect(self.on_search_text_changed)

        self.init_table()

    def on_search_text_changed(self):
        searched_text = self.le_search.text()
        row = 0
        self.table.setRowCount(row)
        if searched_text != "":
            for item in Book_shop.objects():
                if searched_text.upper() in item.name.upper():
                    self.table.setRowCount(row + 1)

                    self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
                    self.table.setItem(row, 1, QTableWidgetItem(item.name))
                    self.table.setItem(row, 2, QTableWidgetItem(str(item.price)))
                    self.table.setItem(row, 3, QTableWidgetItem(str(item.amount)))
                    self.table.setItem(row, 4, QTableWidgetItem(str(item.after_obj.afters_name)))
                    self.table.setItem(row, 5, QTableWidgetItem(str(item.literature_obj.folk_literature)))
                    self.table.setItem(row, 6, QTableWidgetItem(str(item.genre_obj.genre)))
                    self.table.setItem(row, 7, QTableWidgetItem(str(item.publishing_obj.publishing)))

                    row += 1

        else:
            for item in Book_shop.objects():
                self.table.setRowCount(row+1)

                self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
                self.table.setItem(row, 1, QTableWidgetItem(item.name))
                self.table.setItem(row, 2, QTableWidgetItem(str(item.price)))
                self.table.setItem(row, 3, QTableWidgetItem(str(item.amount)))
                self.table.setItem(row, 4, QTableWidgetItem(str(item.after_obj.afters_name)))
                self.table.setItem(row, 5, QTableWidgetItem(str(item.literature_obj.folk_literature)))
                self.table.setItem(row, 6, QTableWidgetItem(str(item.genre_obj.genre)))
                self.table.setItem(row, 7, QTableWidgetItem(str(item.publishing_obj.publishing)))

                row += 1

    def init_table(self):
        #Table
        self.table = QTableWidget(self)
        self.table.setStyleSheet("background-color : white;")
        self.table.setGeometry(500 , 40 , 1000 , 400)

        #Column Count
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["Id" , "Name" , "Price" , "Amount" , "After" , "Literature" , "Genre" , "Publishing"])
        self.table.hideColumn(0)
        self.table.setFont(QFont("Normal" , 14))

        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in Book_shop.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.name))
            self.table.setItem(row, 2, QTableWidgetItem(str(item.price)))
            self.table.setItem(row, 3, QTableWidgetItem(str(item.amount)))
            self.table.setItem(row, 4, QTableWidgetItem(str(item.after_obj.afters_name)))
            self.table.setItem(row, 5, QTableWidgetItem(str(item.literature_obj.folk_literature)))
            self.table.setItem(row, 6, QTableWidgetItem(str(item.genre_obj.genre)))
            self.table.setItem(row, 7, QTableWidgetItem(str(item.publishing_obj.publishing)))

            row += 1

    #Menu
    def init_menu(self):
        menu_bar = self.menuBar()
        file_menu = QMenu("&File", self)
        file_menu.setFont(QFont("Normal", 14))
        menu_bar.addMenu(file_menu)

        service_menu = QMenu("&Services", self)
        service_menu.setFont(QFont("Normal", 14))
        service_menu.setStyleSheet("background-color: white;")
        menu_bar.addMenu(service_menu)

        #Actions
        self.act_after = QAction("&After" , self)
        self.act_after.triggered.connect(self.on_click_btn_act_after)
        service_menu.addAction(self.act_after)

        self.act_literature = QAction("&Literature" , self)
        self.act_literature.triggered.connect(self.on_cilick_btn_act_literature)
        service_menu.addAction(self.act_literature)

        self.act_genre = QAction("&Genre" , self)
        self.act_genre.triggered.connect(self.on_click_btn_act_genre)
        service_menu.addAction(self.act_genre)

        self.act_publishing = QAction("&Publishing" , self)
        self.act_publishing.triggered.connect(self.on_click_btn_act_publishing)
        service_menu.addAction(self.act_publishing)

        self.act_user = QAction("&Users" , self)
        self.act_user.triggered.connect(self.on_click_btn_act_user)
        service_menu.addAction(self.act_user)

    def on_click_btn_act_after(self):
        window = After_window(parent = self)
        window.setWindowTitle("After")
        window.show()

    def on_cilick_btn_act_literature(self):
        window = Literature_window(parent = self)
        window.setWindowTitle("Literature")
        window.show()

    def on_click_btn_act_genre(self):
        window = Genre_window(parent = self)
        window.setWindowTitle("Genre")
        window.show()

    def on_click_btn_act_publishing(self):
        window = Publishing_window(parent = self)
        window.setWindowTitle("Publishing")
        window.show()

    def on_click_btn_act_user(self):
        window = User_window(parent = self)
        window.setWindowTitle("Users")
        window.showMaximized()

    def on_click_table(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()
        self.le_name.setText(self.table.item(row, 1).text())
        self.le_price.setText(self.table.item(row, 2).text())
        self.sb_amount.setValue(int(self.table.item(row , 3).text()))

    def clear_fileds(self):
        self.le_name.setText("")
        self.le_price.setText("")
        self.sb_amount.setValue(1)

    def on_click_btn_report(self):

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        row = 1
        ws[f"A{row}"] = "Name"
        ws[f"B{row}"] = "Price"
        ws[f"C{row}"] = "Amount"
        ws[f"D{row}"] = "After name"
        ws[f"E{row}"] = "Folk Literature"
        ws[f"F{row}"] = "Genre"
        ws[f"G{row}"] = "Publishing"

        row += 1

        for item in Book_shop.objects():
            ws[f"A{row}"] = item.name
            ws[f"B{row}"] = item.price
            ws[f"C{row}"] = item.amount
            ws[f"D{row}"] = item.after
            ws[f"E{row}"] = item.literature
            ws[f"F{row}"] = item.genre
            ws[f"G{row}"] = item.publishing

            row += 1

        file, check = QFileDialog.getSaveFileName(self , "QFileDialog.getOpenFileName()",
                                                  "", "Excel Files (*.xls);")

        if check:
            wb.save(file)
            wb.close()


    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()
        id = int(self.table.item(row, 0).text())

        Book_shop.del_by_id(id)
        self.table.removeRow(row)

    def on_click_btn_update(self):
        try:
            name = self.le_name.text()
            price = int(self.le_price.text())

            if not (name and price):
                raise ValueError ("Maxsulotning nomini yoki narxini kiritmadingiz!")

            select_value = self.table.selectedItems()[0]
            row = select_value.row()
            id = int(self.table.item(row, 0).text())

            amount = int(self.sb_amount.text())
            after = int(self.cb_after.currentData())
            literature = int(self.cb_literature.currentData())
            genre = int(self.cb_genre.currentData())
            publishing = int(self.cb_publishing.currentData())

            new_book = Book_shop(name , price , amount , after, literature, genre, publishing, id)
            new_book.update()

            self.table.setItem(row, 1, QTableWidgetItem(new_book.name))
            self.table.setItem(row, 2, QTableWidgetItem(str(new_book.price)))
            self.table.setItem(row, 3, QTableWidgetItem(str(new_book.amount)))
            self.table.setItem(row, 4, QTableWidgetItem(str(new_book.after_obj.afters_name)))
            self.table.setItem(row, 5, QTableWidgetItem(str(new_book.literature_obj.folk_literature)))
            self.table.setItem(row, 6, QTableWidgetItem(str(new_book.genre_obj.genre)))
            self.table.setItem(row, 7, QTableWidgetItem(str(new_book.publishing_obj.publishing)))

        except ValueError as exc:
            traceback.print_exc()
            self.msg = QMessageBox()
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setInformativeText(str(exc))
            self.msg.setWindowTitle("Kiritishda xatolik!")
            self.msg.setDetailedText("The details are as follows:")
            self.msg.show()

    def on_click_btn_save(self):
        try:
            name = self.le_name.text()
            price = int(self.le_price.text())

            if not (name and price):
                raise ValueError("Maxsulotning nomi yoki narxini kiritmadingiz!")

            amount = int(self.sb_amount.text())
            after = int(self.cb_after.currentData())
            literature = int(self.cb_literature.currentData())
            genre = int(self.cb_genre.currentData())
            publishing = int(self.cb_publishing.currentData())

            new_book = Book_shop(name, price, amount, after, literature, genre, publishing)
            new_book.save()

            row = self.table.rowCount()
            self.table.setRowCount(row + 1)

            self.table.setItem(row , 0 , QTableWidgetItem(str(new_book.id)))
            self.table.setItem(row , 1 , QTableWidgetItem(new_book.name))
            self.table.setItem(row , 2 , QTableWidgetItem(str(new_book.price)))
            self.table.setItem(row , 3 , QTableWidgetItem(str(new_book.amount)))
            self.table.setItem(row , 4 , QTableWidgetItem(str(new_book.after_obj.afters_name)))
            self.table.setItem(row , 5 , QTableWidgetItem(str(new_book.literature_obj.folk_literature)))
            self.table.setItem(row , 6 , QTableWidgetItem(str(new_book.genre_obj.genre)))
            self.table.setItem(row , 7 , QTableWidgetItem(str(new_book.publishing_obj.publishing)))

            self.clear_fileds()
        except ValueError as exc:
            self.msg = QMessageBox()
            self.msg.setIcon(QMessageBox.Critical)
            self.msg.setInformativeText(str(exc))
            self.msg.setWindowTitle("Error⚠️")
            self.msg.setDetailedText(str(exc))
            self.msg.show()

class After_window(QDialog):

    def __init__(self, parent):
        super().__init__(parent)

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setGeometry(100, 100, 800, 600)
        self.setWindowTitle("Afters table")

        #After
        self.lb_after = QLabel(self)
        self.lb_after.setText("After")
        self.lb_after.setFont((QFont("Normal" , 20)))
        self.lb_after.setGeometry(30 , 40 , 200 , 40)

        self.le_after = QLineEdit(self)
        self.le_after.setFont(QFont("Normal" , 14))
        self.le_after.setGeometry(160 , 40 , 200 , 40)
        self.le_after.setStyleSheet("QLineEdit" "{"
                                    "background : white;"
                                    "border : 8px solid;"
                                    "border-top-color : grey;"
                                    "border-left-color : grey;"
                                    "border-right-color : #ebebe0;"
                                    "border-bottom-color : #ebebe0;" "}")

        #Button
        self.btn_save = QPushButton(self)
        self.btn_save.setFont(QFont("Normal", 12))
        self.btn_save.setText("Save")
        self.btn_save.setGeometry(160 , 100 , 200 , 45)
        self.btn_save.clicked.connect(self.on_click_btn_save)
        QBTN = """
                QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.btn_save.setStyleSheet(QBTN)

        self.btn_update = QPushButton(self)
        self.btn_update.setFont(QFont("Normal", 12))
        self.btn_update.setText("Update")
        self.btn_update.setGeometry(160 , 160 , 200 , 45)
        self.btn_update.clicked.connect(self.on_click_btn_update)
        QBTN = """
                QPushButton:active{
                  background-color : #00ff00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                QPushButton:pressed{
                  background-color : #00cc00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                """
        self.btn_update.setStyleSheet(QBTN)

        self.btn_delete = QPushButton(self)
        self.btn_delete.setFont(QFont("Normal", 12))
        self.btn_delete.setText("Delete")
        self.btn_delete.setGeometry(160 , 220 , 200 , 40)
        self.btn_delete.clicked.connect(self.on_click_btn_delete)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.btn_delete.setStyleSheet(QBTN)

        #Table
        self.table = QTableWidget(self)
        self.table.setGeometry(410 , 30 , 230 , 400)
        self.table.setStyleSheet("background-color : white;")

        #Column Count
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Id" , "After name"])
        self.table.hideColumn(0)
        self.table.setColumnWidth(1 , 200)

        self.table.setFont(QFont("Normal" , 14))
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in After.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.afters_name))

            row += 1

    def on_click_btn_save(self):
        name = self.le_after.text()

        after = After(name)
        after.save()

        row = self.table.rowCount()
        self.table.setRowCount(row + 1)

        self.table.setItem(row , 0 , QTableWidgetItem(str(after.id)))
        self.table.setItem(row , 1 , QTableWidgetItem(after.afters_name))

        self.le_after.setText("")

    def on_click_btn_update(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        name = self.le_after.text()
        id = int(self.table.item(row , 0).text())

        after = After(name , id)
        after.update()

        self.table.setItem(row, 1, QTableWidgetItem(after.afters_name))
        self.le_after.setText("")

    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        id = int(self.table.item(row , 0).text())

        After.del_by_id(id)

        self.le_after.setText("")
        self.table.removeRow(row)

    def on_click_table(self):
        select_value =self.table.selectedItems()[0]
        row = select_value.row()
        self.le_after.setText(self.table.item(row, 1).text())


class Literature_window(QDialog):

    def __init__(self, parent):
        super().__init__(parent)

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setWindowTitle("Folk Literature Table")
        self.setGeometry(100, 100, 800, 600)

        #Literature
        self.lb_literature = QLabel(self)
        self.lb_literature.setText("Literature")
        self.lb_literature.setFont(QFont("Normal" , 20))
        self.lb_literature.move(30 , 40)

        self.le_literature = QLineEdit(self)
        self.le_literature.setFont(QFont("Normal" , 14))
        self.le_literature.setGeometry(200 , 40 , 200 , 40)
        self.le_literature.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        #Button
        self.button_save = QPushButton(self)
        self.button_save.setText("Save")
        self.button_save.setFont(QFont("Normal" , 12))
        self.button_save.setGeometry(200 , 100 , 200 , 40)
        self.button_save.clicked.connect(self.on_click_btn_save)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.button_save.setStyleSheet(QBTN)

        self.button_update = QPushButton(self)
        self.button_update.setText("Update")
        self.button_update.setFont(QFont("Normal", 12))
        self.button_update.setGeometry(200, 160, 200, 40)
        self.button_update.clicked.connect(self.on_click_btn_update)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.button_update.setStyleSheet(QBTN)

        self.button_delete = QPushButton(self)
        self.button_delete.setText("Delete")
        self.button_delete.setFont(QFont("Normal", 12))
        self.button_delete.setGeometry(200, 220, 200, 40)
        self.button_delete.clicked.connect(self.on_click_btn_delete)
        QBTN = """
            QPushButton:active{
              background-color : #00ff00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            QPushButton:pressed{
              background-color : #00cc00;
              border: 8px solid #00ff00;
              border-top-color : grey;
              border-left-color : grey;
              border-right-color : #ebebe0;
              border-bottom-color : #ebebe0;
            }
            """
        self.button_delete.setStyleSheet(QBTN)


        #Table
        self.table = QTableWidget(self)
        self.table.setStyleSheet("background-color : white;")
        self.table.setGeometry(480 , 30 , 230 , 400)

        #Column Count
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Id" , "Literature"])
        self.table.hideColumn(0)
        self.table.setColumnWidth(1 , 200)

        self.table.setFont(QFont("Normal" , 14))
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in Literature.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.folk_literature))

            row += 1

    def on_click_btn_save(self):
        name = self.le_literature.text()

        literature = Literature(name)
        literature.save()

        row = self.table.rowCount()
        self.table.setRowCount(row + 1)

        self.table.setItem(row, 0, QTableWidgetItem(str(literature.id)))
        self.table.setItem(row, 1, QTableWidgetItem(literature.folk_literature))

        self.le_literature.setText("")

    def on_click_btn_update(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        name = self.le_literature.text()
        id = int(self.table.item(row, 0).text())

        after = Literature(name, id)
        after.update()

        self.table.setItem(row, 1, QTableWidgetItem(after.folk_literature))
        self.le_literature.setText("")

    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        id = int(self.table.item(row, 0).text())

        Literature.del_by_id(id)

        self.le_literature.setText("")
        self.table.removeRow(row)

    def on_click_table(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()
        self.le_literature.setText(self.table.item(row, 1).text())


class Genre_window(QDialog):

    def __init__(self, parent):
        super().__init__(parent)

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setWindowTitle("Genre Table")
        self.setGeometry(100, 100, 800, 600)

        #Genre
        self.lb_genre = QLabel(self)
        self.lb_genre.setText("Genre")
        self.lb_genre.setFont(QFont("Normal" , 20))
        self.lb_genre.move(30 , 40)

        self.le_genre = QLineEdit(self)
        self.le_genre.setFont(QFont("Normal" , 14))
        self.le_genre.setGeometry(160 , 40 , 200 , 40)
        self.le_genre.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        #Button
        self.button_save = QPushButton(self)
        self.button_save.setText("Save")
        self.button_save.setFont(QFont("Normal" , 12))
        self.button_save.setGeometry(160 , 100 , 200 , 40)
        self.button_save.clicked.connect(self.on_click_btn_save)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_save.setStyleSheet(QBTN)

        self.button_update = QPushButton(self)
        self.button_update.setText("Update")
        self.button_update.setFont(QFont("Normal", 12))
        self.button_update.setGeometry(160, 160, 200, 40)
        self.button_update.clicked.connect(self.on_click_btn_update)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_update.setStyleSheet(QBTN)

        self.button_delete = QPushButton(self)
        self.button_delete.setText("Delete")
        self.button_delete.setFont(QFont("Normal", 12))
        self.button_delete.setGeometry(160, 220, 200, 40)
        self.button_delete.clicked.connect(self.on_click_btn_delete)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_delete.setStyleSheet(QBTN)

        #Table
        self.table = QTableWidget(self)
        self.table.setGeometry(410 , 30 , 230 , 400)
        self.table.setStyleSheet("background-color : white;")

        #Column Count
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Id" , "Genre"])
        self.table.hideColumn(0)
        self.table.setColumnWidth(1 , 200)

        self.table.setFont(QFont("Normal" , 14))
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in Genre.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.genre))

            row += 1

    def on_click_btn_save(self):
        name = self.le_genre.text()

        genre = Genre(name)
        genre.save()

        row = self.table.rowCount()
        self.table.setRowCount(row + 1)

        self.table.setItem(row, 0, QTableWidgetItem(str(genre.id)))
        self.table.setItem(row, 1, QTableWidgetItem(genre.genre))

        self.le_genre.setText("")

    def on_click_btn_update(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        name = self.le_genre.text()
        id = int(self.table.item(row, 0).text())

        genre = Genre(name, id)
        genre.update()

        self.table.setItem(row, 1, QTableWidgetItem(genre.genre))
        self.le_genre.setText("")

    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        id = int(self.table.item(row, 0).text())

        Genre.del_by_id(id)

        self.le_genre.setText("")
        self.table.removeRow(row)

    def on_click_table(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()
        self.le_genre.setText(self.table.item(row, 1).text())


class Publishing_window(QDialog):

    def __init__(self, parent):
        super().__init__(parent)

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setWindowTitle("Publishing Table")
        self.setGeometry(100, 100, 800, 600)

        #Publishing
        self.lb_publishing = QLabel(self)
        self.lb_publishing.setText("Publishing")
        self.lb_publishing.setFont(QFont("Normal" , 20))
        self.lb_publishing.move(30 , 40)

        self.le_publishing = QLineEdit(self)
        self.le_publishing.setFont(QFont("Normal" , 14))
        self.le_publishing.setGeometry(220 , 40 , 200 , 40)
        self.le_publishing.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        #Button
        self.button_save = QPushButton(self)
        self.button_save.setText("Save")
        self.button_save.setFont(QFont("Normal" , 12))
        self.button_save.setGeometry(220 , 100, 200 , 40)
        self.button_save.clicked.connect(self.on_click_btn_save)
        QBTN = """
                QPushButton:active{
                  background-color : #00ff00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                QPushButton:pressed{
                  background-color : #00cc00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                """
        self.button_save.setStyleSheet(QBTN)

        self.button_update = QPushButton(self)
        self.button_update.setText("Update")
        self.button_update.setFont(QFont("Normal", 12))
        self.button_update.setGeometry(220, 160, 200, 40)
        self.button_update.clicked.connect(self.on_click_btn_update)
        QBTN = """
                QPushButton:active{
                  background-color : #00ff00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                QPushButton:pressed{
                  background-color : #00cc00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                """
        self.button_update.setStyleSheet(QBTN)

        self.button_delete = QPushButton(self)
        self.button_delete.setText("Delete")
        self.button_delete.setFont(QFont("Normal", 12))
        self.button_delete.setGeometry(220, 220, 200, 40)
        self.button_delete.clicked.connect(self.on_click_btn_delete)
        QBTN = """
                QPushButton:active{
                  background-color : #00ff00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                QPushButton:pressed{
                  background-color : #00cc00;
                  border: 8px solid #00ff00;
                  border-top-color : grey;
                  border-left-color : grey;
                  border-right-color : #ebebe0;
                  border-bottom-color : #ebebe0;
                }
                """
        self.button_delete.setStyleSheet(QBTN)

        #Table
        self.table = QTableWidget(self)
        self.table.setStyleSheet("background-color: white;")
        self.table.setGeometry(480 , 30 , 230 , 400)

        #Column Count
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Id" , "Publishing"])
        self.table.hideColumn(0)
        self.table.setColumnWidth(1 , 200)

        self.table.setFont(QFont("Normal" , 14))
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in Publishing.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.publishing))

            row += 1

    def on_click_btn_save(self):
        name = self.le_publishing.text()

        publishing = Publishing(name)
        publishing.save()

        row = self.table.rowCount()
        self.table.setRowCount(row + 1)

        self.table.setItem(row, 0, QTableWidgetItem(str(publishing.id)))
        self.table.setItem(row, 1, QTableWidgetItem(publishing.publishing))

        self.le_publishing.setText("")

    def on_click_btn_update(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        name = self.le_publishing.text()
        id = int(self.table.item(row, 0).text())

        publishing = Publishing(name, id)
        publishing.update()

        self.table.setItem(row, 1, QTableWidgetItem(publishing.publishing))
        self.le_publishing.setText("")

    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        id = int(self.table.item(row, 0).text())

        Publishing.del_by_id(id)

        self.le_publishing.setText("")
        self.table.removeRow(row)

    def on_click_table(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()
        self.le_publishing.setText(self.table.item(row, 1).text())


class User_window(QDialog):

    def __init__(self,parent):
        super().__init__(parent)

        self.setWindowIcon(QIcon("book_shop_foto.png"))
        self.setWindowTitle("User Table")
        self.setStyleSheet("background-color: #dd99ff;")
        self.setGeometry(100, 100, 800, 600)

        #Users
        self.lb_first_name = QLabel(self)
        self.lb_first_name.setText("First Name")
        self.lb_first_name.setFont(QFont("Normal" , 20))
        self.lb_first_name.move(30 , 40)

        self.le_first_name = QLineEdit(self)
        self.le_first_name.setFont(QFont("Normal" , 14))
        self.le_first_name.setGeometry(220 , 40 , 200 , 40)
        self.le_first_name.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        self.lb_last_name = QLabel(self)
        self.lb_last_name.setText("Last Name")
        self.lb_last_name.setFont(QFont("Normal" , 20))
        self.lb_last_name.move(30 , 100)

        self.le_last_name = QLineEdit(self)
        self.le_last_name.setFont(QFont("Normal" , 14))
        self.le_last_name.setGeometry(220 , 100 , 200 , 40)
        self.le_last_name.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        self.lb_user_name = QLabel(self)
        self.lb_user_name.setText("User Name")
        self.lb_user_name.setFont(QFont("Normal" , 20))
        self.lb_user_name.move(30 , 160)

        self.le_user_name = QLineEdit(self)
        self.le_user_name.setFont(QFont("Normal" , 14))
        self.le_user_name.setGeometry(220 , 160 , 200 , 40)
        self.le_user_name.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        self.lb_password = QLabel(self)
        self.lb_password.setText("Password")
        self.lb_password.setFont(QFont("Normal" , 20))
        self.lb_password.move(30 , 220)

        self.le_password = QLineEdit(self)
        self.le_password.setFont(QFont("Normal" , 14))
        self.le_password.setGeometry(220 , 220 , 200 , 40)
        self.le_password.setStyleSheet("QLineEdit {"
                                         "background : white;"
                                         "border : 8px solid;"
                                         "border-top-color : grey;"
                                         "border-left-color : grey;"
                                         "border-right-color : #ebebe0;"
                                         "border-bottom-color : #ebebe0;" "}")

        #Button
        self.button_save = QPushButton(self)
        self.button_save.setText("Save")
        self.button_save.setFont(QFont("Normal" , 12))
        self.button_save.setGeometry(220 , 280 , 200 , 40)
        self.button_save.clicked.connect(self.on_click_btn_save)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_save.setStyleSheet(QBTN)

        self.button_update = QPushButton(self)
        self.button_update.setText("Update")
        self.button_update.setFont(QFont("Normal", 12))
        self.button_update.setGeometry(220, 340, 200, 40)
        self.button_update.clicked.connect(self.on_click_btn_update)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_update.setStyleSheet(QBTN)

        self.button_delete = QPushButton(self)
        self.button_delete.setText("Delete")
        self.button_delete.setFont(QFont("Normal", 12))
        self.button_delete.setGeometry(220, 400, 200, 40)
        self.button_delete.clicked.connect(self.on_click_btn_delete)
        QBTN = """
                    QPushButton:active{
                      background-color : #00ff00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    QPushButton:pressed{
                      background-color : #00cc00;
                      border: 8px solid #00ff00;
                      border-top-color : grey;
                      border-left-color : grey;
                      border-right-color : #ebebe0;
                      border-bottom-color : #ebebe0;
                    }
                    """
        self.button_delete.setStyleSheet(QBTN)

        #Table
        self.table = QTableWidget(self)
        self.table.setGeometry(480 , 30 , 1250 , 400)
        self.table.setStyleSheet("background-color : white;")

        #Column Count
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Id" , "First_name", "Last_name" , "User_name" , "Password"])
        self.table.hideColumn(0)
        self.table.setColumnWidth(1 , 200)

        self.table.setFont(QFont("Normal" , 14))
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.clicked.connect(self.on_click_table)

        row = self.table.rowCount()
        for item in User.objects():
            self.table.setRowCount(row + 1)

            self.table.setItem(row, 0, QTableWidgetItem(str(item.id)))
            self.table.setItem(row, 1, QTableWidgetItem(item.first_name))
            self.table.setItem(row, 2, QTableWidgetItem(item.last_name))
            self.table.setItem(row, 3, QTableWidgetItem(item.user_name))
            self.table.setItem(row, 4, QTableWidgetItem(item.password))

            row += 1

    def on_click_btn_save(self):
        first_name = self.le_first_name.text()
        last_name = self.le_last_name.text()
        user_name = self.le_user_name.text()
        password = self.le_password.text()

        user = User(first_name, last_name, user_name)
        user.save()
        user.set_password(password)

        row = self.table.rowCount()
        self.table.setRowCount(row + 1)

        self.table.setItem(row, 0, QTableWidgetItem(str(user.id)))
        self.table.setItem(row, 1, QTableWidgetItem(user.first_name))
        self.table.setItem(row, 2, QTableWidgetItem(user.last_name))
        self.table.setItem(row, 3, QTableWidgetItem(user.user_name))
        self.table.setItem(row, 4, QTableWidgetItem(user.password))

        self.le_first_name.setText("")
        self.le_last_name.setText("")
        self.le_user_name.setText("")
        self.le_password.setText("")

    def on_click_btn_update(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        first_name = self.le_first_name.text()
        last_name = self.le_last_name.text()
        user_name = self.le_user_name.text()
        password_ = self.le_password.text()

        id = int(self.table.item(row, 0).text())

        user = User(first_name, last_name, user_name, password_, id)
        user.update()
        user.set_password(password_)

        self.table.setItem(row, 1, QTableWidgetItem(user.first_name))
        self.table.setItem(row, 2, QTableWidgetItem(user.last_name))
        self.table.setItem(row, 3, QTableWidgetItem(user.user_name))
        self.table.setItem(row, 4, QTableWidgetItem(user.password))

        self.le_first_name.setText("")
        self.le_last_name.setText("")
        self.le_user_name.setText("")
        self.le_password.setText("")

    def on_click_btn_delete(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        id = int(self.table.item(row, 0).text())

        User.del_by_id(id)

        self.le_first_name.setText("")
        self.le_last_name.setText("")
        self.le_user_name.setText("")
        self.le_password.setText("")
        self.table.removeRow(row)

    def on_click_table(self):
        select_value = self.table.selectedItems()[0]
        row = select_value.row()

        self.le_first_name.setText(self.table.item(row, 1).text())
        self.le_last_name.setText(self.table.item(row, 2).text())
        self.le_user_name.setText(self.table.item(row, 3).text())
        self.le_password.setText(self.table.item(row, 4).text())

